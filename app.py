from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)

def buscar_cliente(file_path, cliente):
    wb = load_workbook(file_path)
    sheet = wb.active

    encontrado = False
    cliente_info = []

    for row in sheet.iter_rows(values_only=True):
        if cliente in row:
            encontrado = True
            for col_index, cell_value in enumerate(row, start=1):
                cliente_info.append((f"Coluna {col_index}", cell_value))
            break

    if not encontrado:
        cliente_info.append(("Status", "Cliente não encontrado na planilha."))

    return cliente_info

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file_path = request.form["file_path"]
        cliente = request.form["cliente"]
        cliente_info = buscar_cliente(file_path, cliente)
        return render_template("resultado.html", cliente_info=cliente_info)
    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
